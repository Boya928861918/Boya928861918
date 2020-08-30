import openpyxl
import os

#自动按城市制表
wbs = openpyxl.load_workbook('E:\学习\空气污染\stations.xlsx')
wss = wbs["stationID"]
city = input('请输入城市：')
city = ' '+city
Snumber = []
countcity = 0
for rows in range(1,wss.max_row+1):
    sheetCity = wss.cell(row=rows, column=1).value
    if sheetCity == city:
        print(int(wss.cell(row=rows, column=3).value))
        Snumber.append(int(wss.cell(row=rows, column=3).value))
        countcity += 1

#获取表格数值
def openxlsx(num):
    wb = openpyxl.load_workbook(r'E:\学习\空气污染\广东省AQI数据\2019\AQI_{}A_2019.xlsx'.format(num))
    ws = wb.active

    Numbers_row = []
    Numbers = []
    for colNum in range(2, 8):
        for rowNum in range(2,ws.max_row+1):
            value = ws.cell(row=rowNum, column=colNum).value
            if value == None:
               value  = 0
            Numbers_row.append(value)
        Numbers.append(Numbers_row)
        Numbers_row = []

    times = []
    for rowNum in range(1,ws.max_row+1):
        time = ws.cell(row=rowNum, column=1).value
        times.append(time)

    colN = ws.max_row
    return Numbers,colN,times

print(Snumber)
#获取表格数据
station = []
for num in Snumber:
    r = openxlsx(num)
    station.append(r[0])

colN = r[1]

#创建表格
wb = openpyxl.Workbook()
ws = wb['Sheet']

#处理数据
sites = []
values = 0
count = 0
for colRow in range(0,2160):
    for col in range(0,6):
        for site in range(0,countcity):
            number = station[site][col][colRow]
            values += number
            if number != 0:
                count += 1
        if values == 0:
            continue
        value = values / count
        ws.cell(row=colRow + 2, column=col + 2).value = value
        print(value)
        values = 0
        count = 0


#加入表头
names = ['PM2.5', 'PM10', 'O3', 'CO', 'SO2', 'NO2']
for col in range(2,8):
    ws.cell(row=1, column=col).value = names[col - 2]

for rowNum in range(1,2162):
    ws.cell(row = rowNum, column = 1).value = r[2][rowNum - 1]




wb.save(r'E:\学习\空气污染\城市数据\2019\{}.xlsx'.format(city))